#!/usr/bin/env python
# encoding:utf-8
#__author__ = 'sainter'

import pandas  as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

print ('''然后修改下面的 imput_file文件名称，output_file文件名称，sys_cpu_image和sys_data_image的图片名称
         直接执行python Capacity_sys.py 即可
      ''')

imput_file = '../expfile/20210521/CVM-host-20210521.xlsx'
output_file = '../expfile/20210521/CVM-sys-20210521.xlsx'
sys_cpu_image = '../expfile/20210521/sys_cpu_image.png'
sys_data_image = '../expfile/20210521/sys_data_image.png'

#读取原始数据,记录total_host
df=pd.read_excel(imput_file,usecols="O:T")
total_host = df.shape[0]
print ("总行数为： %d" %total_host)

writer = pd.ExcelWriter(output_file)
sys_number = df.groupby(['physystemID','物理子系统中文名称'])["physystemID"].count().reset_index(name="共有多少台设备")

#计算资源汇总:系统cpu容量评估表
cap_cpu = pd.crosstab(df['physystemID'], df['CPU容量评估结论'])
#cap_cpu.to_excel(excel_writer=writer, sheet_name='系统cpu容量评估表')

#数据盘资源汇总:系统data盘容量评估表
cap_sys = pd.crosstab(df['physystemID'], df['文件系统容量评估结论数据盘'])
#cap_sys.to_excel(excel_writer=writer, sheet_name='系统data盘容量评估表')

#合并容量评估
result_cpu_sys = pd.merge(left=cap_cpu,right=cap_sys,how='inner',left_index=True,right_index=True)
#重置索引
result_cpu_sys = result_cpu_sys.reset_index(drop=False)
result_cpu_sys.rename(columns={"physystemID":"physystemID_result"},inplace=True)

#再次合并容量评估
result_all = pd.concat([sys_number,result_cpu_sys],join='inner',axis=1)
#print (result_all.info())
result_all = result_all.drop(['physystemID_result'],axis=1)

#输出计算资源容量评估结论
def judgement_cpu_cap(result_all):
    if result_all['未取到数据'] == result_all['共有多少台设备']:
        return '未取到数据'
    return  result_all['容量正常_x']/(result_all['共有多少台设备'] - result_all['未取到数据'])

result_all.loc[:,'容量统计结果_计算资源'] = result_all.apply(judgement_cpu_cap,axis=1)

def conclusion_cpu_cap(result_all):
    if result_all['容量统计结果_计算资源'] == '未取到数据':
        return 'N/A'
    if result_all['容量统计结果_计算资源'] <= float(0.6):
        return '系统容量差'
    if result_all['容量统计结果_计算资源'] <= float(0.7) and result_all['容量统计结果_计算资源'] > float(0.6):
        return '系统容量中'
    if result_all['容量统计结果_计算资源'] <= float(0.8) and result_all['容量统计结果_计算资源'] > float(0.7):
        return '系统容量良'
    return '系统容量优'

result_all.loc[:, '容量统计结论_计算资源'] = result_all.apply(conclusion_cpu_cap,axis=1)

#输出数据盘容量评估结论
def judgement_data_cap(result_all):
    if result_all['未挂载数据盘'] == result_all['共有多少台设备']:
        return '未挂载数据盘'
    return  result_all['容量正常_y']/(result_all['共有多少台设备'] - result_all['未挂载数据盘'])

result_all.loc[:,'容量统计结果_数据盘资源'] = result_all.apply(judgement_data_cap,axis=1)

def conclusion_data_cap(result_all):
    if result_all['容量统计结果_数据盘资源'] == '未挂载数据盘':
        return 'N/A'
    if result_all['容量统计结果_数据盘资源'] <= float(0.6):
        return '系统容量差'
    if result_all['容量统计结果_数据盘资源'] <= float(0.7) and result_all['容量统计结果_数据盘资源'] > float(0.6):
        return '系统容量中'
    if result_all['容量统计结果_数据盘资源'] <= float(0.8) and result_all['容量统计结果_数据盘资源'] > float(0.7):
        return '系统容量良'
    return '系统容量优'

result_all.loc[:, '容量统计结论_数据盘资源'] = result_all.apply(conclusion_data_cap,axis=1)
print (result_all)
#写入系统容量评估汇总表sheet页
result_all.to_excel(excel_writer=writer,float_format='%.4f',sheet_name='系统容量评估汇总表', index=False)

#创建新sheet页 写入系统-计算资源图
print ("---------------------------------------------------------")
total_column = result_all.shape[0]
sys_cap_cpu_you = result_all['容量统计结论_计算资源'].value_counts()['系统容量优']
sys_cap_cpu_liang = result_all['容量统计结论_计算资源'].value_counts()['系统容量良']
sys_cap_cpu_zhong = result_all['容量统计结论_计算资源'].value_counts()['系统容量中']
sys_cap_cpu_cha = result_all['容量统计结论_计算资源'].value_counts()['系统容量差']
sys_cap_cpu_na = result_all['容量统计结论_计算资源'].value_counts()['N/A']
sys_cap_cpu_concat = pd.DataFrame({'系统总量':total_column,
                                  '计算资源优':sys_cap_cpu_you,
                                  '计算资源良':sys_cap_cpu_liang,
                                  '计算资源中':sys_cap_cpu_zhong,
                                  '计算资源差':sys_cap_cpu_cha,
                                  'N/A无法统计-未取到数据':sys_cap_cpu_na},
                                  index=[0])
print (sys_cap_cpu_concat)
#sys_cap_col_concat['系统-计算资源图'] = sys_cap_col_concat['系统-计算资源图'].map(lambda x:format(x,'.2%'))
sys_cap_cpu_concat.to_excel(excel_writer=writer,sheet_name='系统-计算资源图', index=False)

#创建新sheet页 写入系统-存储资源图
print ("---------------------------------------------------------")
sys_cap_data_you = result_all['容量统计结论_数据盘资源'].value_counts()['系统容量优']
sys_cap_data_liang = result_all['容量统计结论_数据盘资源'].value_counts()['系统容量良']
sys_cap_data_zhong = result_all['容量统计结论_数据盘资源'].value_counts()['系统容量中']
sys_cap_data_cha = result_all['容量统计结论_数据盘资源'].value_counts()['系统容量差']
sys_cap_data_na = result_all['容量统计结论_数据盘资源'].value_counts()['N/A']
sys_cap_data_concat = pd.DataFrame({'系统总量':total_column,
                                  '存储资源优':sys_cap_data_you,
                                  '存储资源良':sys_cap_data_liang,
                                  '存储资源中':sys_cap_data_zhong,
                                  '存储资源差':sys_cap_data_cha,
                                  'N/A无法统计-未挂载数据盘':sys_cap_data_na},
                                  index=[0])
print (sys_cap_data_concat)
#sys_cap_col_concat['系统-计算资源图'] = sys_cap_col_concat['系统-计算资源图'].map(lambda x:format(x,'.2%'))
sys_cap_data_concat.to_excel(excel_writer=writer,sheet_name='系统-存储资源图', index=False)

writer.save()
writer.close()

##################画图#########################
# 解决坐标轴刻度负号乱码
plt.rcParams['axes.unicode_minus'] = False
# 解决中文乱码问题
plt.rcParams['font.sans-serif'] = ['Simhei']

print ('============================================================')
fig = plt.figure(0)
sys_cpu_image_data = [sys_cap_cpu_you,sys_cap_cpu_liang,sys_cap_cpu_zhong,sys_cap_cpu_cha,sys_cap_cpu_na]
plt.title('系统-计算资源图')
plt.pie(sys_cpu_image_data, labels= sys_cpu_image_data,radius=0.9,labeldistance=0.7, textprops={'fontsize': 12})
plt.axis('equal')
plt.legend( loc = 'lower right',bbox_to_anchor=(1.1, -0.05) ,labels=['计算资源优','计算资源良','计算资源中','计算资源差','N/A无法统计-未取到数据'],fontsize=8, borderaxespad=0.3)
plt.savefig(sys_cpu_image)

sys_cpu_image_insert = Image(sys_cpu_image)
wb = load_workbook(output_file)
ws=wb['系统-计算资源图']
ws.add_image(sys_cpu_image_insert, 'B5')
wb.save(output_file)
plt. close(0)

print ('============================================================')
fig = plt.figure(1)
sys_data_image_data = [sys_cap_data_you,sys_cap_data_liang,sys_cap_data_zhong,sys_cap_data_cha,sys_cap_data_na]
plt.title('系统-存储资源图')
plt.pie(sys_data_image_data, labels= sys_data_image_data,radius=0.9,labeldistance=0.7, textprops={'fontsize': 12})
plt.axis('equal')
plt.legend( loc = 'lower right',bbox_to_anchor=(1.11, -0.05) ,labels=['存储资源优','存储资源良','存储资源中','存储资源差','N/A无法统计-未挂载数据盘'],fontsize=8, borderaxespad=0.3)
plt.savefig(sys_data_image)

sys_data_image_insert = Image(sys_data_image)
wb = load_workbook(output_file)
ws=wb['系统-存储资源图']
ws.add_image(sys_data_image_insert, 'B5')
wb.save(output_file)
plt. close(1)

print ('OK')