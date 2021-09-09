#!/usr/bin/env python
# encoding:utf-8
#__author__ = 'sainter'

import pandas  as pd
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

print ('''使用方法: 修改原始表里面的第一行列名称，把负载数值改为：cpu，数据盘大小改为：data_cap，数据盘使用率改为：data_PCT，机器类型改为：type
         然后修改下面的 imput_file文件名称，output_file文件名称，host_cpu_image和host_data_image的图片名称
         直接执行python Capacity.py 即可
      ''')

imput_file = '../impfile/20210521/CVM信息-20210521.xlsx'
output_file = '../expfile/20210521/CVM-host-20210521.xlsx'
host_cpu_image = '../expfile/20210521/host_cpu_image.png'
host_data_image = '../expfile/20210521/host_data_image.png'

#读取原始数据,记录total_host
df=pd.read_excel(imput_file)
total_host = df.shape[0]
print ("总行数为： %d" %total_host)

#过滤掉type为空的数据
df = df.dropna(axis=0,subset=['type'])
#替换cpu和数据盘为为空的数据，显示为"未挂载数据盘"
df['cpu'] = df['cpu'].fillna(value='未取到数据')
df['data_cap'] = df['data_cap'].fillna(value='未挂载数据盘')
df['used_data_cap'] = df['used_data_cap'].fillna(value='未挂载数据盘')
df['data_PCT'] = df['data_PCT'].fillna(value='未挂载数据盘')

#判断cpu核心数 如果为1，则cpu最小配置显示为是，否则显示否
def get_cpu_number(df):
    if df['cpu_num'] == 1:
        return '是'
    return '否'

df.loc[:,'是否最小配置'] = df.apply(get_cpu_number,axis=1)

def judgement_cpu_cap(df):
    if df['cpu'] == '未取到数据':
        return '未取到数据'
    if df['是否最小配置'] == '否':
        if df['type'] == 'ljap' or df['type'] == 'ljwb' or df['type'] == 'dswb' or df['type'] == 'wb' or df['type'] == 'ap' or df['type'] == 'liap' or df['type'] == 'web代理' or df['type'] == '联机ap' or df['type'] == 'plwb':
            if df['cpu'] <= 0.1:
                return '容量过剩'
            if df['cpu'] > 0.5:
                return '容量不足'
            return '容量正常'
        if df['type'] == 'plap' or df['type'] == '非联机ap':
            if df['cpu'] <= 0.1:
                return '容量过剩'
            if df['cpu'] > 1:
                return '容量不足'
            return '容量正常'
        if df['type'] == 'dsap' or df['type'] == 'ds' or df['type'] == '数据分析ap':
            if df['cpu'] <= 0.2:
                return '容量过剩'
            if df['cpu'] > 0.8:
                return '容量不足'
            return '容量正常'
        if df['type'] == 'dsdb' or df['type'] == 'ljdb'or df['type'] == 'pldb' or df['type'] == 'db-all':
            if df['cpu'] <= 0.1:
                return '容量过剩'
            if df['cpu'] > 0.6:
                return '容量不足'
            return '容量正常'
    return '容量正常'

df.loc[:,'CPU容量评估结论'] = df.apply(judgement_cpu_cap,axis=1)

#复制data_cap列为tmp_data_cap列，并处理掉GB
df.loc[:,'tmp_data_cap'] = df['data_cap']
df['tmp_data_cap']=df['tmp_data_cap'].str.replace('GB','')
df['tmp_data_cap']=df['tmp_data_cap'].str.replace('未挂载数据盘','0')
df['tmp_data_cap']=df['tmp_data_cap'].astype("int")
#print (df.info())
#df = df.loc[:,'tmp_data_cap'] = df['data_cap'].str.replace("GB","").astype('int32')
#df = df.loc[:,'tmp_data_cap'] = df['data_cap'].str.replace("GB","")

#文件系统容量评估结论(数据盘):
def judgement_data_cap(df):
    if df['tmp_data_cap'] == 0:
        return '未挂载数据盘'
    if df['tmp_data_cap'] < 2048:
        if df['data_PCT'] <= 0.1:
            return '容量过剩'
        if df['data_PCT'] > 0.7:
            return '容量不足'
        return '容量正常'
    if df['tmp_data_cap'] > 2048 & df['tmp_data_cap'] <= 5120:
        if df['data_PCT'] <= 0.2:
            return '容量过剩'
        if df['data_PCT'] > 0.8:
            return '容量不足'
        return '容量正常'
    if df['tmp_data_cap'] > 5120 & df['tmp_data_cap'] <= 10240:
        if df['data_PCT'] <= 0.3:
            return '容量过剩'
        if df['data_PCT'] > 0.85:
            return '容量不足'
        return '容量正常'
    if df['tmp_data_cap'] >= 10240:
        if df['data_PCT'] <= 0.35:
            return '容量过剩'
        if df['data_PCT'] > 0.9:
            return '容量不足'
        return '容量正常'

df.loc[:,'文件系统容量评估结论数据盘'] = df.apply(judgement_data_cap,axis=1)

df = df.drop(['tmp_data_cap'], axis=1)
#print("获取到所有的值:\n{0}".format(df))

valid_host = df.shape[0]
print ("过滤掉type为空的数据后，当前总行数为： %d" %valid_host)

#导出整理后的数据
writer = pd.ExcelWriter(output_file)
df.to_excel(excel_writer=writer, sheet_name='主机容量评估汇总表', index=False)
print ("写入主机容量评估汇总表成功")

print ("---------------------------------------------------------")
print ("原始数据主机总量为： %d" %total_host)
print ("纳入本次统计主机总量为： %d" %valid_host)
unvalid_host = total_host - valid_host
print ("未纳入统计的主机数量为： %d" %unvalid_host)
host_cap_pct =  (valid_host/total_host)
print ("主机数据覆盖率为： %f" %host_cap_pct)
print (df['CPU容量评估结论'].value_counts())

if '容量不足' not in df['CPU容量评估结论'].value_counts():
    host_cap_shortage = 0
else:
    host_cap_shortage = df['CPU容量评估结论'].value_counts()['容量不足']
if '容量正常' not in df['CPU容量评估结论'].value_counts():
    host_cap_normal = 0
else:
    host_cap_normal = df['CPU容量评估结论'].value_counts()['容量正常']
if '容量过剩' not in df['CPU容量评估结论'].value_counts():
    host_cap_surplus = 0
else:
    host_cap_surplus = df['CPU容量评估结论'].value_counts()['容量过剩']
if '未取到数据' not in df['CPU容量评估结论'].value_counts():
    host_not_cpudata = 0
else:
    host_not_cpudata = df['CPU容量评估结论'].value_counts()['未取到数据']
#host_cap_shortage = df['CPU容量评估结论'].value_counts()['容量不足']

#创建新sheet页 写入主机-计算资源数据
print ("---------------------------------------------------------")
host_cap_col_concat = pd.DataFrame({'原始表主机总量':total_host,
                                  '纳入统计主机总量':valid_host,
                                  '主机容量正常':host_cap_normal,
                                  '主机容量不足':host_cap_shortage,
                                  '主机容量过剩':host_cap_surplus,
                                  '主机未取到数据': host_not_cpudata,
                                  '主机数据覆盖率': host_cap_pct,},
                                  index=[0])
print (host_cap_col_concat)
host_cap_col_concat['主机数据覆盖率'] = host_cap_col_concat['主机数据覆盖率'].map(lambda x:format(x,'.2%'))

host_cap_col_concat.to_excel(excel_writer=writer,sheet_name='主机-计算资源图', index=False)

#创建新sheet页 写入主机-存储资源（data）数据
print ("---------------------------------------------------------")
print (df['文件系统容量评估结论数据盘'].value_counts())
print ("纳入本次统计主机总量为： %d" %valid_host)
data_cap_nomount = df['文件系统容量评估结论数据盘'].value_counts()['未挂载数据盘']
data_cap_normal = df['文件系统容量评估结论数据盘'].value_counts()['容量正常']
data_cap_shortage = df['文件系统容量评估结论数据盘'].value_counts()['容量不足']
data_cap_surplus = df['文件系统容量评估结论数据盘'].value_counts()['容量过剩']
data_cap_pct = (data_cap_normal+data_cap_shortage+data_cap_surplus)/valid_host

data_cap_col_concat = pd.DataFrame({'统计主机总量': valid_host,
                                    '数据盘容量正常': data_cap_normal,
                                    '数据盘容量不足':data_cap_shortage,
                                    '数据盘容量过剩':data_cap_surplus,
                                    '未挂载数据盘':data_cap_nomount,
                                    '数据盘覆盖率':data_cap_pct},
                                    index=[0])
print (data_cap_col_concat)
data_cap_col_concat['数据盘覆盖率'] = data_cap_col_concat['数据盘覆盖率'].map(lambda x:format(x,'.2%'))
data_cap_col_concat.to_excel(excel_writer=writer,sheet_name='主机-存储资源图', index=False)
writer.save()
writer.close()

##################画图#########################
# 解决坐标轴刻度负号乱码
plt.rcParams['axes.unicode_minus'] = False
# 解决中文乱码问题
plt.rcParams['font.sans-serif'] = ['Simhei']

print ('============================================================')
fig = plt.figure(0)
cpu_image_data = [host_cap_normal,host_cap_shortage,host_cap_surplus,host_not_cpudata]
plt.bar(range(4), cpu_image_data, align= 'center',color='steelblue', alpha = 0.8)
plt.title('主机-计算资源图')
plt.xticks(range(4),['主机容量正常','主机容量不足','主机容量过剩','主机未取到数据'])
for x,y in enumerate(cpu_image_data):
    plt.text(x, y + 30, '%s' % round(y, 1), ha='center')
plt.savefig(host_cpu_image)

host_cpu_image_insert = Image(host_cpu_image)
wb = load_workbook(output_file)
ws=wb['主机-计算资源图']
ws.add_image(host_cpu_image_insert, 'B5')
wb.save(output_file)
plt. close(0)
print ('============================================================')
fig = plt.figure(1)
data_image_data = [data_cap_normal,data_cap_shortage,data_cap_surplus,data_cap_nomount]
plt.bar(range(4), data_image_data, align= 'center',color='steelblue', alpha = 0.8)
plt.title('主机-存储资源图')
plt.xticks(range(4),['数据盘容量正常','数据盘容量不足','数据盘容量过剩','未挂载数据盘'])
for x,y in enumerate(data_image_data):
    plt.text(x, y + 30, '%s' % round(y, 1), ha='center')
plt.savefig(host_data_image)

host_data_image_insert = Image(host_data_image)
wb = load_workbook(output_file)
ws=wb['主机-存储资源图']
ws.add_image(host_data_image_insert, 'B5')
wb.save(output_file)
plt. close(1)

print ('OK')